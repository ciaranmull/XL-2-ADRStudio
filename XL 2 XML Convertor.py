#!/usr/bin/env python
import PySimpleGUI as sg
import os
import sys
from openpyxl import load_workbook
from yattag import Doc, indent

sg.theme('LightGrey6')  # please make your creations colorful

layout = [  [sg.Text('Microsoft Excel to ADRStudio XML', font='AndaleMono 25', justification = 'center')],
            [sg.Text('--------------------------------', font='AndaleMono 25', justification = 'center')],
            [sg.Text( font='AndaleMono 15'), sg.Input(background_color = 'grey',font='AndaleMono 15',justification = 'right'),
             sg.FileBrowse(button_color = ('black','grey'),font='AndaleMono 15')], 
            [sg.OK('Convert', button_color = ('black','green'), font='AndaleMono 20') ],
            [sg.Text('© MullsoftIE 2021', font='AndaleMono 18', justification = 'center')]
            ]

window = sg.Window('',layout, element_justification = 'center')

event, values = window.read()


excel_file = values [0]



wb = load_workbook(excel_file)
ws = wb.worksheets[0]

numOfLines = 300


# Create Yattag doc, tag and text objects
doc, tag, text = Doc().tagtext()

xml_header = '<?xml version="1.0" encoding="UTF-8"?>'
xml_schema = '<!DOCTYPE plist PUBLIC "-//Apple//DTD PLIST 1.0//EN" "http://www.apple.com/DTDs/PropertyList-1.0.dtd">'

doc.asis(xml_header)
doc.asis(xml_schema)

with tag("plist"):
    
    with tag ("dict"):
        with tag("key"):
            text("Dupe Version")
        with tag("string"):
            text()
        with tag("key"):
            text("Reel")
        with tag("string"):
            text()
        with tag("key"):
            text("Script Format")
        with tag("string"):
            text("Default")
        with tag("key"):
            text("Script Frame Rate")
        with tag("string"):
            text("24FPS")
        with tag("key"):
            text("Script Lines")
        with tag('array'):
            count = 0
            # Use ws.max_row for all rows
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
                row = [cell.value for cell in row]
                    #print(row)
                
                if row [1]== None:
                    sg.popup_error('Completed conversion but with the following exception:\nRow ' + str(count) + ' has no value in it', 
                                   background_color = '#FFDF6C', font='AndaleMono 20', no_titlebar=True)
                    #print ('Completed conversion but with the following exception:')
                    #print('Row ' + str(count) + ' has no value in it')
                    break
                
                else:
                    with tag("dict"):
                            with tag("key"):
                                text("Actor")
                            with tag("string"):
                                text()
                            with tag("key"):
                                text("Character")
                            with tag("string"):
                                text(row[1])
                            with tag("key"):
                                text("Clip Name")
                            with tag("string"):
                                text("{0:03}".format(row[0]))
                            with tag("key"):
                                text("Complete")
                            with tag("string"):
                                text("False")
                            with tag("key"):
                                text("Department")    
                            with tag("string"):
                                text()
                            with tag("key"):
                                text("Dialog")
                            with tag("string"):
                                text(row[4])
                                #text()
                            with tag("key"):
                                text("In Time")
                            with tag("string"):
                                text(row[2])
                            with tag("key"):
                                text("Notes")
                            with tag("string"):
                                text()
                            with tag("key"):
                                text("Number of Takes")
                            with tag("string"):
                                text("0")
                            with tag("key"):
                                text("Open Ended")    
                            with tag("string"):
                                text("False")
                            with tag("key"):
                                text("Out Time")
                            with tag("string"):
                                text(row[3])
                                count += 1
                                
##                else:
##                     break

sg.popup_ok('Conversion completed', background_color = '#9DC88D', text_color = 'black', font='AndaleMono 20', no_titlebar=True)
result = indent(
    doc.getvalue(),
    indentation = '    ',
    #indent_text = True
)
excel_file = os.path.splitext(os.path.basename(excel_file))[0]

with open(excel_file +".xml", "w") as f:
    f.write(result)

window.close()
sys.exit()
